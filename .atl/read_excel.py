import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\gabri\Desktop\ARCOR\eecc\balances_ARCOR.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
ws = wb['raw']
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')

for row in range(1, ws.max_row+1):
    vals = []
    for col in range(1, 25):
        v = ws.cell(row,col).value
        vals.append(str(v) if v is not None else '')
    line = ' | '.join(vals)
    print(f'R{row}: {line}')
